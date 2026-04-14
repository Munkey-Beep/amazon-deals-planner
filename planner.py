#!/usr/bin/env python3
"""
AMAZON DEALS PLANNER GENERATOR — Deal Recommendation Template Based
============================================================================
Input:  Amazon Deals Recommendation Template (.xlsx)  +  Fee Preview CSV
Output: Excel workbook with DEALS PLANNER + PEAK CALENDAR

Only reads data from the 'Deal Recommendation Template' tab of the Amazon file.
Fees come from the Amazon Fee Preview report (estimated-fee-total per SKU).
"""

import csv
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═════════════════════════════════════════════════════════════════════════════
# CONSTANTS & HELPERS
# ═════════════════════════════════════════════════════════════════════════════

NAVY       = "1F4E79"
BLUE       = "2E75B6"
LT_BLUE    = "BDD7EE"
YEL        = "FFF2CC"
GRN        = "E2EFDA"
GRN_DARK   = "70AD47"
RED_L      = "FFE0E0"
WHT        = "FFFFFF"
ALT        = "F5F9FF"
GRY        = "F2F2F2"
ORG        = "FCE4D6"
ORG_DARK   = "F4B942"
PURPLE     = "7030A0"

FONT_NAME  = "Calibri"


def col_letter(n):
    return get_column_letter(n)


def fill(c):
    return PatternFill("solid", fgColor=c, start_color=c)


def bdr(color="9DC3E6"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def font(bold=False, color="000000", sz=10, italic=False, name=None):
    return Font(name=name or FONT_NAME, bold=bold, color=color, size=sz, italic=italic)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[col_letter(i)].width = w


# ═════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═════════════════════════════════════════════════════════════════════════════

def clean_header(h):
    """Strip BOM, quotes, non-alphanumeric prefixes from CSV/header strings."""
    h = h.strip().lstrip('\ufeff')
    h = re.sub(r'^[^a-zA-Z]+', '', h)
    return h.strip('"').strip("'").strip()


def load_deal_recommendations(filepath):
    """
    Read the 'Deal Recommendation Template' sheet from Amazon's .xlsx file.
    Headers are in row 4, field-name row in row 5, data starts row 6.
    Schedule is inherited from the first row of each recommendation_id group.
    Returns a list of dicts.
    """
    recs = []
    try:
        wb_src = load_workbook(filepath, data_only=True)

        # Find target sheet
        target = None
        for name in wb_src.sheetnames:
            if "deal recommendation template" in name.lower():
                target = wb_src[name]
                break
        if target is None:
            # Fall back: find any sheet where row 4 has "Deal Type"
            for name in wb_src.sheetnames:
                ws = wb_src[name]
                for col in range(1, 25):
                    v = ws.cell(4, col).value
                    if v and "deal type" in str(v).lower():
                        target = ws
                        break
                if target:
                    break

        if target is None:
            print("⚠️  'Deal Recommendation Template' sheet not found")
            return []

        # Map column positions from row 4 (display headers)
        headers_row4 = []
        for col in range(1, target.max_column + 1):
            v = target.cell(4, col).value
            headers_row4.append(str(v).lower().strip() if v else "")

        def find_col(needle):
            for i, h in enumerate(headers_row4):
                if needle in h:
                    return i  # 0-based
            return -1

        ci = {
            "parent_asin":    find_col("parent asin"),
            "deal_asin":      find_col("deal asin"),
            "product_name":   find_col("product name"),
            "deal_type":      find_col("deal type"),
            "rec_id":         find_col("recommendation id"),
            "sku":            find_col("sku"),
            "participating":  find_col("participating"),
            "schedule":       find_col("schedule"),
            "seller_price":   find_col("seller price"),
            "deal_price":     find_col("deal price"),
            "committed":      find_col("committed units"),
            "seller_qty":     find_col("seller quantity"),
        }

        # Track schedule inheritance by recommendation_id
        rec_schedule_map = {}

        for row in range(6, target.max_row + 1):
            def cv(key):
                idx = ci.get(key, -1)
                if idx < 0:
                    return None
                return target.cell(row, idx + 1).value

            sku = str(cv("sku") or "").strip()
            if not sku:
                continue

            deal_type = str(cv("deal_type") or "").strip()
            rec_id    = str(cv("rec_id")    or "").strip()

            # Schedule: only real strings, not formula strings
            raw_sched = cv("schedule")
            if raw_sched and not str(raw_sched).startswith("="):
                schedule = str(raw_sched).strip()
                if rec_id and rec_id not in rec_schedule_map:
                    rec_schedule_map[rec_id] = schedule
            else:
                schedule = rec_schedule_map.get(rec_id, "")

            def to_float(v):
                try:
                    return float(v) if v not in (None, "", "--") else 0.0
                except (ValueError, TypeError):
                    return 0.0

            def to_int(v):
                try:
                    return int(float(v)) if v not in (None, "", "--") else 0
                except (ValueError, TypeError):
                    return 0

            recs.append({
                "parent_asin":    str(cv("parent_asin") or "").strip(),
                "deal_asin":      str(cv("deal_asin")   or "").strip(),
                "product_name":   str(cv("product_name") or "").strip(),
                "deal_type":      deal_type,
                "recommendation_id": rec_id,
                "sku":            sku,
                "participating":  str(cv("participating") or "Yes").strip(),
                "schedule":       schedule,
                "seller_price":   to_float(cv("seller_price")),
                "deal_price":     to_float(cv("deal_price")),
                "committed_units": to_int(cv("committed")),
                "seller_quantity": to_int(cv("seller_qty")),
            })

        print(f"✓ Loaded {len(recs)} deal recommendations from '{target.title}'")
        return recs

    except Exception as e:
        print(f"❌ Error loading deal recommendations: {e}")
        import traceback; traceback.print_exc()
        return []


def load_fees(filepath):
    """
    Load Amazon Fee Preview CSV.
    Prefers 'estimated-fee-total' (referral + fulfillment combined).
    Falls back to 'expected-fulfillment-fee-per-unit' if total not available.
    Filters to US store rows only when multiple marketplaces are present.
    Returns two dicts: (sku_map, asin_map) — both keyed to estimated total fee.
    For backward compatibility, also returns a merged flat dict as .combined attribute.
    Use lookup_fee(sku, asin, sku_map, asin_map) for best-effort matching.
    """
    sku_map  = {}   # sku  -> fee
    asin_map = {}   # asin -> fee
    if not filepath:
        return sku_map, asin_map
    try:
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                rows_all = []
                with open(filepath, encoding=enc) as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        cleaned = {
                            clean_header(k): v.strip().strip('"') if v else ""
                            for k, v in row.items() if k
                        }
                        rows_all.append(cleaned)

                # Prefer US-only rows if file contains multiple marketplaces
                us_rows = [r for r in rows_all if r.get("amazon-store", "US") == "US"]
                target_rows = us_rows if us_rows else rows_all

                for cleaned in target_rows:
                    sku  = cleaned.get("sku",  "").strip()
                    asin = cleaned.get("asin", "").strip()

                    fee_str = cleaned.get("estimated-fee-total", "").strip()
                    if not fee_str or fee_str == "--":
                        fee_str = cleaned.get("expected-fulfillment-fee-per-unit", "0").strip()

                    try:
                        fee = float(fee_str) if fee_str and fee_str != "--" else 0.0
                    except ValueError:
                        fee = 0.0

                    if sku:
                        sku_map[sku] = fee
                    if asin:
                        asin_map[asin] = fee

                if sku_map or asin_map:
                    break
            except (UnicodeDecodeError, UnicodeError):
                continue
    except Exception as e:
        print(f"⚠️  Warning: Could not load fees file: {e}")

    print(f"  Fee Preview: {len(sku_map)} SKUs, {len(asin_map)} ASINs loaded (US store)")
    return sku_map, asin_map


def lookup_fee(sku, asin, sku_map, asin_map):
    """
    Look up the estimated total fee for a SKU.
    Tries SKU first, then ASIN as fallback.
    Returns (fee, source) where source is 'sku', 'asin', or 'missing'.
    """
    if sku and sku in sku_map:
        return sku_map[sku], "sku"
    if asin and asin in asin_map:
        return asin_map[asin], "asin"
    return 0.0, "missing"


# ═════════════════════════════════════════════════════════════════════════════
# FEE / SCHEDULE HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def is_prime_day(schedule):
    return bool(schedule and "prime day" in str(schedule).lower())


def parse_days_from_schedule(schedule):
    """Extract duration in days from 'Mon (2026-04-20 - 2026-04-26)' patterns."""
    if not schedule:
        return 7
    m = re.search(r'(\d{4}-\d{2}-\d{2})\s*[-–]\s*(\d{4}-\d{2}-\d{2})', str(schedule))
    if m:
        try:
            d1 = datetime.strptime(m.group(1), "%Y-%m-%d")
            d2 = datetime.strptime(m.group(2), "%Y-%m-%d")
            return max(1, (d2 - d1).days + 1)
        except Exception:
            pass
    return 7


def compute_upfront_fee(deal_type, schedule):
    """One-time upfront fee per deal submission."""
    if is_prime_day(schedule):
        return 100.0
    if deal_type == "Lightning Deal":
        return 70.0          # single slot (≤12 hrs)
    else:                    # Best Deal
        days = parse_days_from_schedule(schedule)
        return 70.0 * days


def var_fee_rate(schedule):
    return 0.015 if is_prime_day(schedule) else 0.01


def var_fee_cap(schedule):
    return 5000.0 if is_prime_day(schedule) else 2000.0


# ═════════════════════════════════════════════════════════════════════════════
# WORKBOOK CREATION
# ═════════════════════════════════════════════════════════════════════════════

def create_workbook(brand_name, recommendations, fees_data):
    """
    Build the Amazon Deals Planner Excel workbook.
    fees_data: either (sku_map, asin_map) tuple from load_fees(), or a plain dict (legacy).
    Sheets: DEALS PLANNER, PEAK CALENDAR
    """
    # Unpack fees_data — support both new tuple form and legacy dict form
    if isinstance(fees_data, tuple):
        sku_map, asin_map = fees_data
    else:
        sku_map  = fees_data  # legacy: plain dict
        asin_map = {}

    wb = Workbook()

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1 — DEALS PLANNER
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "DEALS PLANNER"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    NUM_COLS = 17   # A through Q
    last_col  = col_letter(NUM_COLS)

    # ── Row 1: Main title ──────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    c = ws.cell(1, 1, f"🛒  DEALS PLANNER — {brand_name}")
    c.font      = Font(name=FONT_NAME, bold=True, color=WHT, size=16)
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Row 2: Subtitle ───────────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    c = ws.cell(2, 1,
        "★ All deals are Amazon-recommended  |  "
        "Source: Amazon Deals Recommendation Template  |  "
        "Fees: Amazon Fee Preview (Referral + FBA)")
    c.font      = Font(name=FONT_NAME, italic=True, color=WHT, size=10)
    c.fill      = fill(BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Row 3: COGS note ──────────────────────────────────────────────────
    ws.merge_cells(f"A3:{last_col}3")
    c = ws.cell(3, 1,
        "  ★ = Amazon Recommended Deal  |  "
        "COGS / Unit (column N) is optional — fill yellow cells to calculate true net profit & margin.  |  "
        "🟠 Orange Var Fee = cap prorated across deal group  |  "
        "🔴 Red row = SKU missing from Fee Preview")
    c.font      = Font(name=FONT_NAME, italic=True, color="595959", size=9)
    c.fill      = fill(YEL)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 16

    # ── Row 4: Column Headers ─────────────────────────────────────────────
    HEADERS = [
        # (label, background)
        ("SKU",                         LT_BLUE),
        ("Product Name",                LT_BLUE),
        ("Deal Type ★",                 LT_BLUE),
        ("Schedule / Period",           LT_BLUE),
        ("List Price ($)",              LT_BLUE),
        ("Deal Price ($)",              LT_BLUE),
        ("Disc %",                      LT_BLUE),
        ("Committed Units",             LT_BLUE),
        ("Deal Revenue ($)",            BLUE),
        ("Est. Amazon Fee / Unit ($)",  NAVY),
        ("Deal Var Fee / Unit ($)",     NAVY),
        ("Total Fees / Unit ($)",       NAVY),
        ("Total Fees ($)",              NAVY),
        ("COGS / Unit* ($)",            YEL),
        ("Total COGS* ($)",             YEL),
        ("SKU Profit* ($)",             GRN_DARK),
        ("Margin*",                     GRN_DARK),
    ]
    ws.row_dimensions[4].height = 38
    for col_i, (hdr, bg) in enumerate(HEADERS, 1):
        c = ws.cell(4, col_i, hdr)
        txt_col = WHT if bg in (NAVY, BLUE, GRN_DARK) else "1F1F1F"
        c.font      = Font(name=FONT_NAME, bold=True, color=txt_col, size=9)
        c.fill      = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = bdr()

    # ── Group recommendations by recommendation_id (preserve order) ───────
    rec_groups  = {}   # rid -> [rec, ...]
    rec_order   = []
    for rec in recommendations:
        rid = rec["recommendation_id"]
        if rid not in rec_groups:
            rec_groups[rid]  = []
            rec_order.append(rid)
        rec_groups[rid].append(rec)

    ordered_recs = []
    for rid in rec_order:
        ordered_recs.extend(rec_groups[rid])

    # ── Pre-compute effective variable-fee rate per group (cap-adjusted) ──
    # The cap ($2,000 non-peak / $5,000 Prime Day) applies to the TOTAL
    # variable fees for the whole deal group.  When the cap is hit we prorate:
    #   effective_rate = cap / group_total_revenue
    # This ensures row-level Var Fee/Unit and Total Fees are consistent with
    # the capped figure shown in the Deal Summary.
    group_effective_rates = {}   # rid -> effective rate (float)
    group_cap_hit         = {}   # rid -> bool
    for rid in rec_order:
        group   = rec_groups[rid]
        sched   = group[0]["schedule"]
        raw_rate = var_fee_rate(sched)
        cap      = var_fee_cap(sched)
        grp_rev  = sum(r["deal_price"] * r["committed_units"] for r in group)
        raw_var  = grp_rev * raw_rate
        if grp_rev > 0 and raw_var > cap:
            group_effective_rates[rid] = cap / grp_rev   # prorated
            group_cap_hit[rid]         = True
        else:
            group_effective_rates[rid] = raw_rate
            group_cap_hit[rid]         = False

    # ── Data Rows ─────────────────────────────────────────────────────────
    DATA_START = 5
    row = DATA_START

    for i, rec in enumerate(ordered_recs):
        r        = row
        sku      = rec["sku"]
        schedule = rec["schedule"]
        dtype    = rec["deal_type"]
        sp       = rec["seller_price"]
        dp       = rec["deal_price"]
        units    = rec["committed_units"]
        rid      = rec["recommendation_id"]
        amz_fee, fee_src = lookup_fee(sku, rec["deal_asin"], sku_map, asin_map)
        rate     = group_effective_rates[rid]   # cap-adjusted effective rate
        fee_missing = (fee_src == "missing")

        row_bg = ALT if i % 2 == 0 else WHT
        fee_bg = RED_L if fee_missing else row_bg   # highlight rows with no fee data

        def wc(col_i, value, fmt=None, bold=False, bg_ov=None, align="center", italic=False):
            c = ws.cell(r, col_i, value)
            c.font      = Font(name=FONT_NAME, bold=bold, italic=italic,
                               color="000000", size=9)
            c.fill      = fill(bg_ov if bg_ov is not None else row_bg)
            c.alignment = Alignment(horizontal=align, vertical="center",
                                    wrap_text=(col_i == 2))
            c.border    = bdr()
            if fmt:
                c.number_format = fmt
            return c

        wc(1,  sku,                 align="left")
        wc(2,  rec["product_name"][:90],  align="left")
        wc(3,  dtype)
        wc(4,  schedule or "Non-Peak")
        wc(5,  sp,                  "$#,##0.00")
        wc(6,  dp,                  "$#,##0.00")
        wc(7,  f"=IFERROR((E{r}-F{r})/E{r},0)", "0%")
        wc(8,  units,               "#,##0")
        wc(9,  f"=F{r}*H{r}",      "$#,##0.00")                          # Deal Revenue
        # Est. Amazon Fee/Unit — red background if fee data not found
        c10 = wc(10, amz_fee, "$#,##0.00", bg_ov=fee_bg)                # Est. Amazon Fee/Unit
        if fee_missing:
            c10.font = Font(name=FONT_NAME, size=9, italic=True, color="CC0000")
        elif fee_src == "asin":
            # Matched via ASIN fallback — note in italic
            c10.font = Font(name=FONT_NAME, size=9, italic=True, color="595959")
        # Deal Var Fee/Unit — uses cap-adjusted effective rate when group cap is hit
        c11 = wc(11, f"=F{r}*{rate:.8f}", "$#,##0.00")                   # Deal Var Fee/Unit
        if group_cap_hit[rid]:
            # Faint orange tint to indicate cap was applied; tooltip via cell comment not
            # supported in openpyxl easily, but color signals the proration
            c11.fill = fill(ORG)
            c11.font = Font(name=FONT_NAME, size=9, italic=True, color="7F3F00")
        wc(12, f"=J{r}+K{r}",      "$#,##0.00")                          # Total Fees/Unit
        wc(13, f"=L{r}*H{r}",      "$#,##0.00")                          # Total Fees
        wc(14, "",                  "$#,##0.00", bg_ov=YEL)              # COGS/Unit (user fills)
        wc(15, f"=IF(N{r}>0,N{r}*H{r},0)", "$#,##0.00", bg_ov=YEL)     # Total COGS
        wc(16, f"=I{r}-M{r}-O{r}", "$#,##0.00", bg_ov=GRN,
           bold=True)                                                      # SKU Profit*
        wc(17, f"=IFERROR(P{r}/I{r},0)", "0.0%", bg_ov=GRN)             # Margin*

        ws.row_dimensions[r].height = 20
        row += 1

    DATA_END = row - 1

    # ── Gap before summary ────────────────────────────────────────────────
    row += 1
    SUMMARY_TITLE_ROW = row

    # ── Missing-fee warning row ───────────────────────────────────────────
    missing_skus = [
        rec["sku"] for rec in ordered_recs
        if lookup_fee(rec["sku"], rec["deal_asin"], sku_map, asin_map)[1] == "missing"
    ]
    if missing_skus:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws.cell(row, 1,
            f"⚠️  {len(missing_skus)} SKU(s) have no fee data in your Fee Preview file: "
            f"{', '.join(missing_skus[:6])}{'…' if len(missing_skus) > 6 else ''}  "
            f"— Download a complete Fee Preview from Seller Central → Reports → Fulfillment → Fee Preview "
            f"and re-generate for accurate totals. Rows highlighted in red above.")
        c.font      = Font(name=FONT_NAME, bold=True, color="7F0000", size=9)
        c.fill      = fill("FFE0E0")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 30
        row += 1

    # ── DEAL SUMMARY title bar ────────────────────────────────────────────
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row, 1, "📊  DEAL SUMMARY — Upfront fee charged ONCE per deal group (not per SKU)")
    c.font      = Font(name=FONT_NAME, bold=True, color=WHT, size=12)
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    row += 1

    # Helper: write a single summary row
    def s_row(label, value, fmt="$#,##0.00", note="", label_bold=False,
              val_bold=False, bg=GRY, note_bg=None):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row, 1, f"  {label}")
        c.font      = Font(name=FONT_NAME, bold=label_bold, color="1F1F1F", size=10)
        c.fill      = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(f"F{row}:I{row}")
        c = ws.cell(row, 6, value)
        c.font      = Font(name=FONT_NAME, bold=val_bold, color="1F1F1F", size=10)
        c.number_format = fmt
        c.fill      = fill(bg)
        c.alignment = Alignment(horizontal="right", vertical="center")

        if note:
            ws.merge_cells(f"J{row}:{last_col}{row}")
            c = ws.cell(row, 10, f"  {note}")
            c.font      = Font(name=FONT_NAME, italic=True, color="595959", size=8)
            c.fill      = fill(note_bg or bg)
            c.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[row].height = 18

    # ── Per-group breakdown ───────────────────────────────────────────────
    # Pre-compute row mapping: sku+rec_id → data row number
    data_row_map = {}
    for i, rec in enumerate(ordered_recs):
        data_row_map[(rec["sku"], rec["recommendation_id"])] = DATA_START + i

    for g_idx, rid in enumerate(rec_order):
        group   = rec_groups[rid]
        first   = group[0]
        sched   = first["schedule"]
        dtype   = first["deal_type"]
        prime   = is_prime_day(sched)
        raw_rate = var_fee_rate(sched)
        cap     = var_fee_cap(sched)
        upfront = compute_upfront_fee(dtype, sched)

        # Row numbers in the DEALS PLANNER sheet for this group's SKUs
        grp_rows = [data_row_map[(r["sku"], rid)] for r in group]

        # Excel formulas that reference actual data rows — summary always matches rows
        rev_formula = "=SUM(" + ",".join(f"I{dr}" for dr in grp_rows) + ")"
        var_formula = "=" + "+".join(f"K{dr}*H{dr}" for dr in grp_rows)
        amz_formula = "=" + "+".join(f"J{dr}*H{dr}" for dr in grp_rows)

        # Python values only used for notes / cap display
        grp_revenue   = sum(r["deal_price"] * r["committed_units"] for r in group)
        grp_raw_var   = sum(r["deal_price"] * r["committed_units"] * raw_rate for r in group)
        grp_var_capped = min(grp_raw_var, cap)

        # Group header row
        period_tag = "🔥 PRIME DAY" if prime else "📅 Non-Peak"
        bg_grp = ORG if prime else LT_BLUE

        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws.cell(row, 1,
            f"  Deal Group {g_idx + 1}:  {dtype}  │  {sched or 'Non-Peak'}  │  {period_tag}  │  {len(group)} SKU(s)")
        c.font      = Font(name=FONT_NAME, bold=True, color="1F1F1F", size=10)
        c.fill      = fill(bg_grp)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        s_row("Deal Revenue", rev_formula, bg=GRY, note=f"{len(group)} SKU(s)")
        row += 1

        upfront_note = (
            "$100 flat — Prime Day" if prime else
            (f"$70 flat — Lightning Deal (single slot)" if dtype == "Lightning Deal"
             else f"${upfront:,.0f}  =  {parse_days_from_schedule(sched)} days × $70/day")
        )
        s_row("(Less)  Upfront Fee — ONE-TIME per deal", upfront,
              note=upfront_note, bg=RED_L)
        row += 1

        var_note = (
            f"{raw_rate*100:.1f}% of deal price  |  "
            f"Cap ${cap:,.0f}"
            + (f"  |  Raw ${grp_raw_var:,.2f}  ✓ CAPPED" if grp_raw_var > cap else "")
        )
        s_row("(Less)  Variable Deal Fees (capped)", var_formula,
              note=var_note, bg=RED_L)
        row += 1

        s_row("(Less)  Est. Amazon Fees  (Referral + FBA)", amz_formula,
              note="From Amazon Fee Preview — referral + FBA fulfillment combined", bg=RED_L)
        row += 1

        cogs_formula = f"=SUM({','.join(f'O{dr}' for dr in grp_rows)})"
        s_row("(Less)  Total COGS*", cogs_formula,
              note="Fill COGS/Unit (yellow, column N) in the rows above", bg=YEL)
        row += 1

        # Group sub-profit — formula referencing the summary value cells above
        # summary rows were written at: row-6 (revenue), row-5 (upfront), row-4 (var),
        # row-3 (amz), row-2 (cogs), so profit = revenue - upfront - var - amz - cogs
        rev_row    = row - 6
        upfr_row   = row - 5
        var_row    = row - 4
        amz_row    = row - 3
        cogs_row   = row - 2
        grp_profit_formula = (
            f"=F{rev_row}-F{upfr_row}-F{var_row}-F{amz_row}-F{cogs_row}"
        )
        grp_profit_ex = grp_revenue - upfront - grp_var_capped  # python fallback for display
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row, 1, "  → Est. Group Profit*  (before COGS)")
        c.font      = Font(name=FONT_NAME, bold=True, color="1F4E79", size=10)
        c.fill      = fill(GRN)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(f"F{row}:I{row}")
        c = ws.cell(row, 6, grp_profit_formula)
        c.font         = Font(name=FONT_NAME, bold=True, color="1F4E79", size=10)
        c.number_format = "$#,##0.00"
        c.fill          = fill(GRN)
        c.alignment     = Alignment(horizontal="right", vertical="center")

        ws.merge_cells(f"J{row}:{last_col}{row}")
        c = ws.cell(row, 10, "  * Subtract COGS (fill column N above) for true net profit")
        c.font      = Font(name=FONT_NAME, italic=True, color="595959", size=8)
        c.fill      = fill(GRN)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 2   # blank separator between groups

    # ── GRAND TOTALS ──────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row, 1, "💰  GRAND TOTALS — All Deals Combined")
    c.font      = Font(name=FONT_NAME, bold=True, color=WHT, size=12)
    c.fill      = fill(BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    row += 1

    # Grand Totals — upfront fee is the only Python-only value (no column for it in rows)
    total_upfront = sum(
        compute_upfront_fee(rec_groups[rid][0]["deal_type"], rec_groups[rid][0]["schedule"])
        for rid in rec_order
    )

    def g_row(label, value, fmt="$#,##0.00", note="", bold=False, bg=GRY):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row, 1, f"  {label}")
        c.font      = Font(name=FONT_NAME, bold=bold, color="1F1F1F", size=10)
        c.fill      = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(f"F{row}:I{row}")
        c = ws.cell(row, 6, value)
        c.font         = Font(name=FONT_NAME, bold=bold, color="1F1F1F", size=10)
        c.number_format = fmt
        c.fill          = fill(bg)
        c.alignment     = Alignment(horizontal="right", vertical="center")

        if note:
            ws.merge_cells(f"J{row}:{last_col}{row}")
            c = ws.cell(row, 10, f"  {note}")
            c.font      = Font(name=FONT_NAME, italic=True, color="595959", size=8)
            c.fill      = fill(bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18

    # All totals derived from the data rows via Excel formulas — guaranteed to match
    rev_total_row = row
    g_row("Total Deal Revenue",
          f"=SUM(I{DATA_START}:I{DATA_END})", bold=True,
          note=f"{len(ordered_recs)} SKU(s) across {len(rec_order)} deal group(s)")
    row += 1

    upfr_total_row = row
    g_row("(Less)  Total Upfront Fees",
          -total_upfront, bg=RED_L,
          note=f"{len(rec_order)} deal group(s) — one upfront fee per group")
    row += 1

    var_total_row = row
    g_row("(Less)  Total Variable Deal Fees (capped per group)",
          f"=-SUMPRODUCT(K{DATA_START}:K{DATA_END},H{DATA_START}:H{DATA_END})",
          bg=RED_L,
          note="Each group capped separately ($2K Non-Peak / $5K Prime Day)")
    row += 1

    amz_total_row = row
    g_row("(Less)  Total Est. Amazon Fees  (Referral + FBA)",
          f"=-SUMPRODUCT(J{DATA_START}:J{DATA_END},H{DATA_START}:H{DATA_END})",
          bg=RED_L,
          note="Referral + fulfillment from Amazon Fee Preview")
    row += 1

    cogs_total_row = row
    g_row("(Less)  Total COGS*",
          f"=-SUM(O{DATA_START}:O{DATA_END})",
          note="Fill COGS/Unit (yellow column N) in data rows above", bg=YEL)
    row += 1

    # NET PROFIT bar — formula sums all the g_row cells above
    net_profit_formula = (
        f"=F{rev_total_row}"
        f"+F{upfr_total_row}"
        f"+F{var_total_row}"
        f"+F{amz_total_row}"
        f"+F{cogs_total_row}"
    )

    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row, 1, "  💰  NET DEAL PROFIT*")
    c.font      = Font(name=FONT_NAME, bold=True, color=WHT, size=13)
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells(f"F{row}:I{row}")
    c = ws.cell(row, 6, net_profit_formula)
    c.font         = Font(name=FONT_NAME, bold=True, color=WHT, size=13)
    c.number_format = "$#,##0.00"
    c.fill          = fill(NAVY)
    c.alignment     = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(f"J{row}:{last_col}{row}")
    c = ws.cell(row, 10, "  * Before COGS — fill column N above for true net profit")
    c.font      = Font(name=FONT_NAME, italic=True, color="9DC3E6", size=9)
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    # Net Margin — derived from the NET PROFIT and revenue formula cells above
    net_margin_row = row - 1   # net profit was written to row-1
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row, 1, "  Net Margin*  (before COGS)")
    c.font      = Font(name=FONT_NAME, bold=True, color=WHT, size=11)
    c.fill      = fill(BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells(f"F{row}:I{row}")
    c = ws.cell(row, 6, f"=IFERROR(F{net_margin_row}/F{rev_total_row},0)")
    c.font         = Font(name=FONT_NAME, bold=True, color=WHT, size=11)
    c.number_format = "0.0%"
    c.fill          = fill(BLUE)
    c.alignment     = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 2

    # HOW IT WORKS note
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row, 1,
        "ℹ️  HOW IT WORKS:  All deals are sourced from Amazon's Deals Recommendation Template (Amazon-vetted). "
        "Upfront fee = one charge per deal group submission (not per SKU). "
        "Variable fee = Deal Price × rate per unit (1% Non-Peak / 1.5% Prime Day), capped at $2,000 or $5,000 per group. "
        "Est. Amazon Fee = referral + FBA fulfillment combined (from Fee Preview). "
        "COGS/Unit is optional — fill yellow column N to see true net profit.")
    c.font      = Font(name=FONT_NAME, italic=True, color="595959", size=8)
    c.fill      = fill(GRY)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 36

    # ── Column widths ─────────────────────────────────────────────────────
    set_col_widths(ws, [18, 42, 14, 30, 12, 12, 8, 12, 14, 14, 14, 14, 14, 12, 12, 14, 10])

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 2 — PEAK CALENDAR
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("PEAK CALENDAR")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:H1")
    c = ws2.cell(1, 1, f"📆  PEAK EVENT CALENDAR — {brand_name} 2025/2026")
    c.font      = Font(name=FONT_NAME, bold=True, color=WHT, size=13)
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells("A2:H2")
    c = ws2.cell(2, 1,
        "Planning guide for peak events with fee structures and submission deadlines. "
        "Customize strategy notes for your brand.")
    c.font      = Font(name=FONT_NAME, italic=True, color="555555", size=9)
    c.fill      = fill(GRY)
    ws2.row_dimensions[2].height = 18

    cal_hdrs = [
        ("Event", NAVY), ("Dates", BLUE), ("Best Deal Fee", NAVY),
        ("Lightning Fee", NAVY), ("Coupon Fee", BLUE), ("Min Discount", BLUE),
        ("Strategy Notes", BLUE), ("Brand Opportunity", NAVY),
    ]
    for i, (h, bg) in enumerate(cal_hdrs, 1):
        c = ws2.cell(3, i, h)
        c.font      = Font(name=FONT_NAME, bold=True,
                           color=WHT if bg == NAVY else "1F1F1F", size=10)
        c.fill      = fill(bg)
        c.border    = bdr()
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws2.row_dimensions[3].height = 28

    events = [
        ("Prime Day", "~July (est.)", 100, 100, "$5+2.5%", 0.15,
         "Submit by Apr 30 to save $50/deal. $100 flat upfront + 1.5% variable (capped $5K). "
         "Same fee for both LD and BD.",
         "🔥 HIGH — Peak summer shopping."),
        ("Black Friday / BFCM", "~Late November", "TBD", "TBD", 245, 0.15,
         "Submit 6+ weeks early. Biggest traffic day. Variable fee ~1.5% capped $5K. "
         "Upfront TBD — check Seller Central.",
         "🔥 HIGHEST — Top gift-giving event."),
        ("Prime Big Deal Days", "~October (est.)", "TBD", "TBD", "$5+2.5%", 0.15,
         "Submit 4+ weeks ahead. Variable fee ~1.5% capped $5K. Upfront TBD.",
         "🟡 MEDIUM — Pre-holiday push."),
        ("Cyber Monday", "~December 2", 1000, 500, 245, 0.15,
         "Span Best Deal across BFCM + Cyber Monday for single $1,000 fee. Layer coupons.",
         "🔥 HIGH — Online-focused buyers."),
        ("Non-Peak Deals", "Year-round", "$70/day", "$70 flat", "$5+2.5%", 0.15,
         "Best ROI for margins. Run 7–14 day Best Deals at $70/day. 1% variable, capped $2K.",
         "🟡 MEDIUM — Consistent, predictable fees."),
    ]

    fmts = [None, None, "$#,##0", "$#,##0", "$#,##0", "0%", None, None]
    for idx, evt_row in enumerate(events):
        r = idx + 4
        bg = WHT if idx % 2 == 0 else ALT
        for ci, (val, fmt) in enumerate(zip(evt_row, fmts), 1):
            c = ws2.cell(r, ci, val)
            c.font      = Font(name=FONT_NAME, size=9, bold=(ci in [1, 8]))
            c.fill      = fill(bg)
            c.border    = bdr()
            c.alignment = Alignment(
                horizontal="left" if ci in [1, 6, 7] else "center",
                wrap_text=True)
            if fmt and isinstance(val, (int, float)):
                c.number_format = fmt
        ws2.row_dimensions[r].height = 48

    set_col_widths(ws2, [24, 18, 12, 12, 14, 12, 46, 36])

    print(f"✓ Workbook generated: {len(ordered_recs)} SKUs | {len(rec_order)} deal group(s)")
    return wb


# ═════════════════════════════════════════════════════════════════════════════
# MAIN (CLI)
# ═════════════════════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Generate Amazon Deals Planner from Deals Recommendation Template")
    parser.add_argument("--brand",    required=True,
                        help="Brand name (e.g. 'Grillbot')")
    parser.add_argument("--deals",    required=True,
                        help="Path to Amazon Deals Recommendation Template (.xlsx)")
    parser.add_argument("--fees",     required=False, default=None,
                        help="Path to Amazon Fee Preview CSV (optional)")
    parser.add_argument("--output",   required=True,
                        help="Output .xlsx path")
    args = parser.parse_args()

    print(f"\n  Amazon Deals Planner — {args.brand}")
    print("  Loading deal recommendations...")
    recs = load_deal_recommendations(args.deals)
    if not recs:
        print("❌  No recommendations found. Check the file and sheet name.")
        sys.exit(1)

    fees_data = ({}, {})
    if args.fees:
        print("  Loading fee preview...")
        fees_data = load_fees(args.fees)

    wb = create_workbook(args.brand, recs, fees_data)
    wb.save(args.output)
    print(f"\n  ✅  Saved: {args.output}\n")


if __name__ == "__main__":
    main()
